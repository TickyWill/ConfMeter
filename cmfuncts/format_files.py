"""Module of functions for formatting files as openpyxl workbooks 
may be used by several modules of package `cmfuncts`.
"""


__all__ = ['add_sheets_to_workbook',
           'format_hal_page',
          ]

# 3rd party imports
import pandas as pd
from bmfuncts.format_files import build_cell_fill_patterns
from bmfuncts.format_files import color_row
from bmfuncts.format_files import align_cell
from bmfuncts.format_files import format_heading
from bmfuncts.format_files import set_col_width
from openpyxl import Workbook as openpyxl_Workbook
from openpyxl.utils.dataframe import dataframe_to_rows \
    as openpyxl_dataframe_to_rows
from openpyxl.utils import get_column_letter \
    as openpyxl_get_column_letter
from openpyxl.styles import Font as openpyxl_Font
from openpyxl.styles import PatternFill as openpyxl_PatternFill
from openpyxl.styles import Alignment as openpyxl_Alignment
from openpyxl.styles import Border as openpyxl_Border
from openpyxl.styles import Side as openpyxl_Side

# Local imports
import cmfuncts.conf_globals as cg


def add_sheets_to_workbook(file_full_path, df_to_add, sheet_name):
    """Adds the dataframe 'df_to_add' as sheet named 'sheet_name' 
    to the existing Excel file with full path 'file_full_path'. 

    If the sheet name already exists it is overwritten by the new one.

    Args:
        file_full_path (path): The full path to the file to be completed.
        df_to_add (dataframe): The data for filling the added sheet.
        sheet_name (str): The name of the added sheet.
    """
    with pd.ExcelWriter(file_full_path,  # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
                        engine='openpyxl',
                        mode='a',
                        if_sheet_exists='replace') as writer:
        df_to_add.to_excel(writer, sheet_name=sheet_name, index=False)


def _set_hal_col_attr(cols_rename_dict):
    """Sets the dict for setting the final column attributes 
    in terms of width and alignment to be used for formating 
    data before openpyxl save.

    A default attributes are defined for column names of data 
    not among the keys of the built dict of attributes. 
    The default attributes are given at key 'else' of this dict.

    Args:
        cols_rename_dict (dict): The dict for using the renamed \
        columns of the data.
    Returns:
        (tup): (the dict to be used for setting the columns \
        attributes for formating data before openpyxl save, \
        list of the column names that have attributes).
    """
    init_col_attr = {cg.HASH_COL['hash_id']           : [15, "center"],
                     cg.CONF_COLS['pub_id']           : [15, "center"],
                     cg.CONF_COLS['pub_year']         : [12, "center"],
                     cg.CONF_COLS['conf_year']        : [12, "center"],
                     cg.CONF_COLS['conf_date']        : [12, "center"],
                     cg.CONF_COLS['first_author']     : [25, "center"],
                     cg.CONF_ADD_COLS['inst_authors'] : [40, "left"],                              
                     cg.CONF_COLS['authors']          : [40, "left"], 
                     cg.CONF_COLS['title']            : [40, "left"],
                     cg.CONF_COLS['conf_name']        : [40, "left"],
                     cg.CONF_COLS['doctype']          : [15, "center"],
                     cg.CONF_COLS['doi']              : [20, "center"],
                     cg.CONF_ADD_COLS['full_ref']     : [55, 'left'],
                     cg.CONF_COLS['town']             : [20, "center"],
                     cg.CONF_COLS['country']          : [15, "center"],
                     cg.CONF_COLS['commitee']         : [12, "center"],
                     cg.CONF_COLS['proceedings']      : [12, "center"],
                     cg.CONF_COLS['url']              : [50, "left"]
                    }

    final_cols = [cols_rename_dict[col] for col in init_col_attr.keys()]
    col_attr = dict(zip(final_cols, init_col_attr.values()))
    col_set_list = list(col_attr.keys())
    col_attr['else'] = [10, "center"]

    return col_attr, col_set_list


def format_hal_page(df, cols_rename_dict, wb=None):
    """Formats a worksheet of an openpyxl workbook using 
    columns attributes set through the `_set_hal_col_attr` 
    internal function.

    When the workbook wb is not None, this is applied 
    to the active worksheet of the passed workbook. 
    If the workbook wb is None, then the workbook is created.

    Args:
        df (dataframe): The dataframe to be formatted.
        cols_rename_dict (dict): The dict for using the renamed \
        columns of the data.
        wb (openpyxl workbook): Optional worbook of the worksheet \
        to be formatted (default = None).
    Returns:
        (tup): (worbook of the formatted worksheet (openpyxl workbook), \
        formatted active sheet).
    """
    # Setting useful aliases
    xl_idx_base_alias = cg.XL_INDEX_BASE
    df_title_alias = cg.CONF_DF_TITLE

    # Setting first column to format
    col_idx_init = 0

    # Setting useful column sizes
    col_attr, col_set_list = _set_hal_col_attr(cols_rename_dict)
    columns_list = list(df.columns)
    for col in columns_list:
        if col not in col_set_list:
            col_attr[col] = col_attr['else']

    # Setting list of cell colors
    cell_colors = build_cell_fill_patterns()

    # Initialize wb as a workbook and ws its active worksheet
    if not wb:
        wb = openpyxl_Workbook()
    ws = wb.active

    # Coloring alternately rows in ws
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=True)
    for idx_row, row in enumerate(ws_rows):
        ws.append(row)
        ws = color_row(ws, idx_row, cell_colors)

    # Setting cell alignment and border in ws
    ws = align_cell(ws, columns_list, col_attr, xl_idx_base_alias)

    # Setting the format of the columns heading
    ws = format_heading(ws, df_title_alias)

    # Setting the columns width
    ws = set_col_width(ws, columns_list, col_attr,
                       col_idx_init, xl_idx_base_alias)

    # Setting height of first row
    first_row_num = 1
    ws.row_dimensions[first_row_num].height = 50

    return wb, ws
